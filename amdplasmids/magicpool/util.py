"""
    Various utility functions
"""
import os
import sys
import gzip
import openpyxl
import hashlib
from pathlib import Path
from collections import defaultdict
from Bio import SeqIO
from Bio import GenBank
from subprocess import Popen, PIPE, CalledProcessError
from magicpool.models import *
from amdplasmids.settings import DATA_DIR

def autovivify(levels=1, final=dict):
    return (defaultdict(final) if levels < 2 else
            defaultdict(lambda: autovivify(levels - 1, final)))

def guess_feature_type(name, type_label):
    feature_type_dict = {item.name:item for item in Feature_type.objects.all()}
    if type_label=='misc_feature':
        if name.endswith('romoter'):
            if 'promoter' in feature_type_dict:
                return feature_type_dict['promoter']
            else:
                return Feature_type.objects.create(name='promoter')
        elif 'origin of replication' in name:
            if 'rep_origin' in feature_type_dict:
                return feature_type_dict['rep_origin']
            else:
                return Feature_type.objects.create(name='rep_origin')
        elif name.startswith('Ori'):
            if 'rep_origin' in feature_type_dict:
                return feature_type_dict['rep_origin']
            else:
                return Feature_type.objects.create(name='rep_origin')
        elif name == 'ColE1':
            if 'rep_origin' in feature_type_dict:
                return feature_type_dict['rep_origin']
            else:
                return Feature_type.objects.create(name='rep_origin')
        elif 'BsmBI' in name:
            if 'restriction_site' in feature_type_dict:
                return feature_type_dict['restriction_site']
            else:
                return Feature_type.objects.create(name='restriction_site')
        elif 'inverted repeat' in name:
            if 'IR' in feature_type_dict:
                return feature_type_dict['IR']
            else:
                return Feature_type.objects.create(name='IR')
        elif 'transposase enzyme' in name:
            if 'gene' in feature_type_dict:
                return feature_type_dict['gene']
            else:
                return Feature_type.objects.create(name='gene')
        else:
            if 'misc_feature' in feature_type_dict:
                return feature_type_dict['misc_feature']
            else:
                return Feature_type.objects.create(name='misc_feature')
    elif type_label in feature_type_dict:
        return feature_type_dict[type_label]
    elif type_label == 'CDS':
            if 'CDS' in feature_type_dict:
                return feature_type_dict['CDS']
            else:
                return Feature_type.objects.create(name='CDS')
    else:
        return Feature_type.objects.create(name=type_label)

def create_feature(feature, plasmid_obj):
    feature_type_label = feature.type

    feature_start = int(feature.location.start)
    feature_end = int(feature.location.end)
    feature_strand = int(feature.location.strand)
    if Feature.objects.filter(plasmid = plasmid_obj,
            start = feature_start,
            end = feature_end,
            strand = feature_strand
            ).exists():
        # Feature exists. Skip it.
        return 0

    if feature_type_label == 'source':
        return 0
    feature_name = ''
    if 'locus_tag' in feature.qualifiers:
        feature_name = str(feature.qualifiers['locus_tag'][0])
    elif 'label' in feature.qualifiers:
        feature_name = str(feature.qualifiers['label'][0])
    elif 'name' in feature.qualifiers:
        feature_name = str(feature.qualifiers['name'][0])
    if feature_name == '':
        print('This feature has no name:')
        print(str(feature))
        return 0
    feature_name = feature_name.replace('\n','')
    description = ''
    if 'gene' in feature.qualifiers:
        description += str(';' .join(feature.qualifiers['gene']))
    if 'label' in feature.qualifiers:
        if description != '':
            description += '; '
        description += str(';' .join(feature.qualifiers['label']))
    if 'note' in feature.qualifiers:
        if description != '':
            description += '; '
        description += str(feature.qualifiers['note'][0])
    description = description.replace('\n','')
    feature_type = guess_feature_type(feature_name, feature_type_label)
    feature_obj = Feature.objects.create(
        name = feature_name,
        plasmid = plasmid_obj,
        feature_type = feature_type,
        sequence = plasmid_obj.sequence[feature_start:feature_end],
        sequence_id = plasmid_obj.name,
        start = feature_start,
        end = feature_end,
        strand = feature_strand,
        location_str = str(feature.location),
        description = description
        )
    if feature_type.name == 'gene' or 'translation' in feature.qualifiers:
        if 'gene' in feature.qualifiers:
            protein_name = str(feature.qualifiers['gene'][0])
        elif 'label' in feature.qualifiers:
            protein_name = str(feature.qualifiers['label'][0])
        elif 'locus_tag' in feature.qualifiers:
            protein_name = str(feature.qualifiers['locus_tag'][0])
        else:
            protein_name = 'unknown_protein'
        protein_name = protein_name.replace('\n','')
        protein_name = protein_name.replace('&nbsp;',' ')
        if 'product' in feature.qualifiers:
            protein_function = str(feature.qualifiers['product'][0])
        elif 'note' in feature.qualifiers:
            protein_function = str(feature.qualifiers['note'][0])
        elif 'label' in feature.qualifiers:
            protein_function = str(feature.qualifiers['label'][0])
        else:
            protein_function = 'unidentified function'

        if 'translation' in feature.qualifiers:
            protein_sequence = ''.join(feature.qualifiers['translation'])
        else:
            protein_sequence = ''
            
        protein_obj = Protein.objects.create(
            name = protein_name,
            sequence = protein_sequence,
            function = protein_function,
            feature = feature_obj
        )
    return 1    

    
def update_plasmid(plasmid, seq_record, sequence_file, checksum):
    # Replace sequence, wipe out old features and create new features
    ret_val = 0
    feature_count = 0
    plasmid.sequence = str(seq_record.seq)
    plasmid.sequence_file = sequence_file
    plasmid.footprint = checksum
    plasmid.save()
    Feature.objects.filter(plasmid=plasmid.id).delete()
    Protein.objects.filter(feature=None).delete()
    
    for feature in seq_record.features:
        feature_count += create_feature(feature, plasmid)
    print(plasmid.name, 'object updated')
    ret_val = 1
    return ret_val, feature_count
    

def import_seq_records(records, seq_name, sequence_file, checksum, overwrite_existing=False):
    record_count = 0
    for seq_record in records:
        record_count += 1
        if record_count > 1:
            continue
        existing_plasmids = Plasmid.objects.filter(name=seq_name)
        feature_count = 0
        if existing_plasmids.exists():
            # Check sequence
            existing_plasmid = existing_plasmids.first()
            if overwrite_existing:
                if existing_plasmid.footprint == checksum and existing_plasmid.sequence_file == sequence_file:
                    # Nothing to do here
                    print(existing_plasmid.name, 'did not change')
                    ret_val = 0
                elif existing_plasmid.sequence_file == sequence_file:
                    # The file has changed. Update plasmid object
                    ret_val, feature_count = update_plasmid(existing_plasmid, seq_record, sequence_file, checksum)
                else:
                    print('New file found for', existing_plasmid.name)
                    if existing_plasmid.sequence_file == '':
                        ret_val, feature_count = update_plasmid(existing_plasmid, seq_record, sequence_file, checksum)
                    elif existing_plasmid.sequence_file.endswith('.fa') or existing_plasmid.sequence_file.endswith('.fna'):
                        ret_val, feature_count = update_plasmid(existing_plasmid, seq_record, sequence_file, checksum)
                    elif existing_plasmid.sequence_file.endswith('.gb') or existing_plasmid.sequence_file.endswith('.ape'):
                        if sequence_file.endswith('.dna') or sequence_file.endswith('.gb') or sequence_file.endswith('.ape'):
                            ret_val, feature_count = update_plasmid(existing_plasmid, seq_record, sequence_file, checksum)
                        else:
                            print(sequence_file, 'was skipped because Genbank format is preferred')
                            ret_val = 0
                    elif existing_plasmid.sequence_file.endswith('.dna'):
                        if sequence_file.endswith('.dna'):
                            ret_val, feature_count = update_plasmid(existing_plasmid, seq_record, sequence_file, checksum)
                        else:
                            print(sequence_file, 'was skipped because Snapgene format is preferred')
                            ret_val = 0
                    else:
                        print('File format is not supported:', sequence_file)
                        ret_val = 0
            else:
                print(existing_plasmid.name, 'object not updated because the --overwrite option is not set')
                ret_val = 0
        else:
            # Create new plasmid object and features
            plasmid_obj = Plasmid.objects.create(
                name = seq_name,
                amd_number = '',
                description = '',
                sequence = str(seq_record.seq),
                footprint = checksum,
                sequence_file = sequence_file
            )
            for feature in seq_record.features:
                feature_count += create_feature(feature, plasmid_obj)
            print(seq_name, ' object created')
            ret_val = 1
    if record_count > 1:
        print(record_count, 'records provided for import. Only the first record will be imported')
    print(feature_count, 'features created')
    return ret_val

def import_plasmid_snapgene(dna_file, sequence_file, overwrite_existing):
    print('Working on Snapgene file', dna_file)
    with open(dna_file, 'rb') as infile:
        checksum = hashlib.md5(infile.read()).hexdigest()
    records = SeqIO.parse(dna_file, "snapgene")
    filename = dna_file.split('/')[-1]
    try:
        ret_val = import_seq_records(records, filename.split('.dna')[0], sequence_file, checksum, overwrite_existing)
    except:
        print('Snapgene file conversion error', dna_file)
        ret_val = 0
    return ret_val
            
def import_plasmid_fasta(fna_file, sequence_file, overwrite_existing):
    print('Working on FASTA file', fna_file)
    with open(fna_file, 'rb') as infile:
        checksum = hashlib.md5(infile.read()).hexdigest()
    records = SeqIO.parse(fna_file, "fasta")
    filename = fna_file.split('/')[-1]
    try:
        ret_val = import_seq_records(records, '.'.join(filename.split('.')[:-1]), sequence_file, checksum, overwrite_existing)
    except:
        print('FASTA file conversion error', fna_file)
        ret_val = 0
    return ret_val
            
def import_plasmid_gbk(gbk_file, sequence_file, overwrite_existing):
    print('Working on GenBank file', gbk_file)
    if gbk_file.endswith('.gz'):
        gbk_handle = gzip.open(gbk_file, 'rb')
    else:
        gbk_handle = open(gbk_file, 'rb')
    checksum = hashlib.md5(gbk_handle.read()).hexdigest()
    gbk_handle.close()

    if gbk_file.endswith('.gz'):
        gbk_handle = gzip.open(gbk_file, 'rt')
    else:
        gbk_handle = open(gbk_file, 'r')
    records = SeqIO.parse(gbk_handle, "genbank")
    filename = gbk_file.split('/')[-1]
    ret_val = import_seq_records(records, '.'.join(filename.split('.')[:-1]), sequence_file, checksum, overwrite_existing)
    gbk_handle.close()
    return ret_val


def update_alldata(overwrite=False):
    if not os.path.exists(DATA_DIR):
        print("Data directory does not exists: " + DATA_DIR)
        return
    else:
        print("Data directory: " + DATA_DIR)
    plasmid_maps_dir = os.path.join(DATA_DIR, 'plasmid_maps')
    oligos_file = os.path.join(DATA_DIR, 'oligos.xlsx')
    strains_file = os.path.join(DATA_DIR, 'strains.xlsx')
    # Download files
    if not os.path.exists(os.path.join(DATA_DIR, 'rclone.conf')):
        raise FileNotFoundError('rclone.conf is missing')
    with open(os.path.join(DATA_DIR, 'data_download.sh'), 'w') as outfile:
        outfile.write('#!/usr/bin/bash\n')
        outfile.write('cd '+ DATA_DIR + '\n')
        outfile.write('rclone sync --config ./rclone.conf "gdriveR:Deutschbauer_Lab_Documents/PLASMID_MAPS_2023_onward" --drive-shared-with-me ./plasmid_maps\n')
        outfile.write('rclone copyto --config ./rclone.conf "gdriveR:Deutschbauer_Lab_Documents/AMD strain collection.xlsx" --drive-shared-with-me ./strains.xlsx\n')
        outfile.write('rclone copyto --config ./rclone.conf "gdriveR:Deutschbauer_Lab_Documents/Oligos_and_gBlocks/oAD_oligos_and_gAD_gBlocks.xlsx" --drive-shared-with-me ./oligos.xlsx\n')
    cmd = ['bash', os.path.join(DATA_DIR, 'data_download.sh')]
    print(' '.join(cmd))
    with Popen(cmd, stdout=PIPE, bufsize=1, universal_newlines=True) as proc:
        for line in proc.stdout:
            print(line.rstrip('\n\r'))
    if proc.returncode != 0:
        # Suppress false positive no-member error
        # (see https://github.com/PyCQA/pylint/issues/1860)
        # pylint: disable=no-member
        raise CalledProcessError(proc.returncode, proc.args)

    # Import magic pool types and vector designs before importing plasmid files
    import_magic_pool_types(os.path.join(plasmid_maps_dir, 'magicpool_vector_designs', 'magic_pool_design.xlsx'))
        
    if not os.path.exists(plasmid_maps_dir):
        print("Data directory does not exists: " + plasmid_maps_dir)
        return
    else:
        print("Plasmid data directory found: " + plasmid_maps_dir)
    existing_plasmids_nomap = Plasmid.objects.filter(sequence_file='')
    delete_objects = []
    for plasmid in Plasmid.objects.exclude(sequence_file=''):
        if not os.path.exists(os.path.join(plasmid_maps_dir, plasmid.sequence_file)):
            print('File not found: ' + os.path.join(plasmid_maps_dir, plasmid.sequence_file))
            print(plasmid.name + ' plasmid to be deleted')
            delete_objects.append(plasmid)
    if delete_objects:
        for item in delete_objects:
            item.delete()
        delete_objects = []

    plasmids_nomap = import_plasmids(plasmid_maps_dir, overwrite)
    for plasmid_obj in existing_plasmids_nomap:
        if plasmid_obj.name not in plasmids_nomap:
            print(plasmid.name + ' plasmid to be deleted')
            delete_objects.append(plasmid)
    if delete_objects:
        for item in delete_objects:
            item.delete()
        delete_objects = []
    
    import_magic_pools(os.path.join(plasmid_maps_dir, 'Magic_Pools', 'Magic_Pool_Summary_Sheet.xlsx'))

    existing_oligos = Oligo.objects.all()
    new_oligo_names = import_oligos_table(oligos_file)
    delete_objects = []
    for oligo_obj in existing_oligos:
        if oligo_obj.name not in new_oligo_names:
            print(oligo_obj.name + ' oligo to be deleted')
            delete_objects.append(oligo_obj)
    if delete_objects:
        for item in delete_objects:
            item.delete()
        delete_objects = []
    
    existing_strains = Strain.objects.all()
    new_strain_amd_numbers = import_strains_table(strains_file)
    delete_objects = []
    for strain_obj in existing_strains:
        if strain_obj.amd_number not in new_strain_amd_numbers:
            print(strain_obj.amd_number + ' strain to be deleted')
            delete_objects.append(strain_obj)
    if delete_objects:
        for item in delete_objects:
            item.delete()
    make_blast_databases()


def import_plasmids(work_dir, overwrite_existing):
    obj_count = 0
    ret = []
    for root, subdirs, files in os.walk(work_dir):
        if 'Old_files' in root or 'Older_Files' in root:
            print('Skipping', root)
            continue
        print(files)
        for filename in sorted(files):
            filepath = os.path.join(root, filename)
            if filename.endswith('.dna'):
                obj_count += import_plasmid_snapgene(filepath, filepath.split(work_dir)[-1][1:], overwrite_existing)
            elif filename.endswith('.gb') or filename.endswith('.ape'):
                obj_count += import_plasmid_gbk(filepath, filepath.split(work_dir)[-1][1:], overwrite_existing)
            elif filename.endswith('.xlsx'):
                if filename.startswith('Old_'):
                    continue
                ret = ret + import_plasmids_table(filepath, overwrite_existing)
            elif filename.endswith('.fa') or filename.endswith('.fna'):
                obj_count += import_plasmid_fasta(filepath, filepath.split(work_dir)[-1][1:], overwrite_existing)
    print(obj_count, 'plasmids created and/or updated')
    return ret

    
def create_plasmid(plasmid_name, plasmid_data):
    amd_number = ''
    description = ''
    sequence = ''
    footprint = hashlib.md5(sequence.encode('utf-8')).hexdigest()
    if 'AMD number' in plasmid_data:
        amd_number = plasmid_data['AMD number']
    if 'Description' in plasmid_data:
        description = plasmid_data['Description']
    plasmid_obj = Plasmid.objects.create(
        name = plasmid_name,
        amd_number = amd_number,
        description = description,
        sequence = sequence,
        footprint = footprint
    )
    if 'Magic pool part number' in plasmid_data:
        plasmid_obj.magic_pool_designation = plasmid_data['Magic pool part number']
        plasmid_obj.save()

    if 'Magic pool part type' in plasmid_data:
        existing_parts = {item.name:item for item in Magic_pool_part_type.objects.all()}
        part_name = plasmid_data['Magic pool part type']
        if part_name in existing_parts:
            plasmid_obj.magic_pool_part_type = existing_parts[part_name]
            plasmid_obj.save()
        else:
            part_obj = Magic_pool_part_type.objects.create(
                name = part_name,
                description = '',
                upstream_overhang = None,
                downstream_overhang = None,
            )
            plasmid_obj.magic_pool_part = part_obj
            plasmid_obj.save()
        
    if 'Drug marker' in plasmid_data:
        for marker in plasmid_data['Drug marker'].split('; '):
            try:
                marker_obj = Drug_marker.objects.get(name=marker)
            except Drug_marker.DoesNotExist:
                marker_obj = Drug_marker.objects.create(
                    name = marker,
                    drug = marker,
                    note = ''
                    )
            plasmid_obj.drug_markers.add(marker_obj)
            plasmid_obj.save()
    for key, value in plasmid_data.items():
        if key in ('AMD number', 'Description', 'Drug marker', 'Magic pool part type', 'Magic pool part number'):
            continue
        plasmid_info_obj = Plasmid_info.objects.create(
            plasmid = plasmid_obj,
            param = key,
            value = value
        )
    return 1
    
def import_plasmids_table(xlsx_path, overwrite_existing=False):
    print(xlsx_path)
    ret = []
    xlsx_path = Path(xlsx_path)
    wb_obj = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb_obj.active
    xlsx_header = []
    data_imported = defaultdict(dict)
    created_count = 0
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            xlsx_header = row[1:]
        else:   
            plasmid_name = row[0]
            if plasmid_name == '' or plasmid_name is None:
                continue
            ret.append(plasmid_name)
            for j, cell in enumerate(row[1:]):
                if cell != '' and cell != 'None' and cell is not None:
                    data_imported[plasmid_name][xlsx_header[j]] = str(cell)
    if 'AMD number' not in xlsx_header:
        print(created_count, 'new plasmids created')
        return ret
    for plasmid_name, plasmid_data in data_imported.items():
        if Plasmid.objects.filter(name=plasmid_name).exists():
            plasmid = Plasmid.objects.get(name=plasmid_name)
            if not overwrite_existing:
                continue
            for key,value in plasmid_data.items():
                if key == 'AMD number':
                    if plasmid.amd_number != value:
                        plasmid.amd_number = value
                        plasmid.save()
                elif key == 'Drug marker':
                    existing_markers = [item.name for item in plasmid.drug_markers.all()]
                    for marker in value.split('; '):
                        if marker not in existing_markers:
                            try:
                                marker_obj = Drug_marker.objects.get(name=marker)
                            except Drug_marker.DoesNotExist:
                                marker_obj = Drug_marker.objects.create(
                                    name = marker,
                                    drug = marker,
                                    note = ''
                                    )
                            plasmid.drug_markers.add(marker_obj)
                            plasmid.save()
                elif key == 'Description':
                    if value not in plasmid.description:
                        plasmid.description += value
                        plasmid.save()
                elif key == 'Magic pool part type':
                    existing_parts = {item.name:item for item in Magic_pool_part_type.objects.all()}
                    if value in existing_parts:
                        plasmid.magic_pool_part = existing_parts[value]
                        plasmid.save()
                    else:
                        part_obj = Magic_pool_part_type.objects.create(
                            name = value,
                            description = '',
                            upstream_overhang = None,
                            downstream_overhang = None,
                        )
                        plasmid.magic_pool_part = part_obj
                        plasmid.save()
                elif key == 'Magic pool part number':
                    if value != plasmid.magic_pool_designation:
                        plasmid.magic_pool_designation = value
                        plasmid.save()
                else:
                    if Plasmid_info.objects.filter(plasmid=plasmid.id,param=key).exists():
                        plasmid_info_obj = Plasmid_info.objects.get(plasmid=plasmid.id,param=key)
                        if plasmid_info_obj.value != value:
                            plasmid_info_obj.value = value
                            plasmid_info_obj.save()
                    else:
                        Plasmid_info.objects.create(
                            plasmid = plasmid,
                            param = key,
                            value = value
                            )
        else:
            created_count += create_plasmid(plasmid_name, plasmid_data)
    print(created_count, 'new plasmids created')
    return ret

def import_magic_pool_types(xlsx_path):
    xlsx_path = Path(xlsx_path)
    wb_obj = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb_obj.get_sheet_by_name('part_names_overlaps')
    xlsx_header = []
    created_count = 0
    for i, row in enumerate(sheet.iter_rows()):
        if i == 0:
            xlsx_header = [cell.value for cell in row[1:]]
        else:
            existing_parts = {item.name:item for item in Magic_pool_part_type.objects.all()}
            part_name = row[0].value
            if part_name == '' or part_name is None:
                continue
            if part_name in existing_parts:
                continue
            existing_overhangs = {item.name:item for item in Overhang.objects.all()}
            description = ''
            upstream_overhang = None
            upstream_overhang_color = None
            downstream_overhang = None
            downstream_overhang_color = None
            partinfo = []
            for j, cell in enumerate(row[1:]):
                if xlsx_header[j] == '' or xlsx_header[j] is None:
                    break
                cell_text = cell.value
                if cell_text != '' and cell_text != 'None' and cell_text is not None:
                    if xlsx_header[j] == 'Contains':
                        description = cell_text
                    elif xlsx_header[j] == '4bp overhang upstream':
                        upstream_overhang = cell_text
                        cell_color = int(cell.fill.start_color.index, 16)
                        upstream_overhang_color = "%06x" % (cell_color and 0xFFFFFF)
                    elif xlsx_header[j] == '4bp overhang downstream':
                        downstream_overhang = cell_text
                        cell_color = int(cell.fill.start_color.index, 16)
                        downstream_overhang_color = "%06x" % (cell_color and 0xFFFFFF)
                    else:
                        partinfo.append((xlsx_header[j], cell_text))
            if upstream_overhang is None:
                print(part_name, ': Upstream overhang not found')
                continue
            if downstream_overhang is None:
                print(part_name, ': Downstream overhang not found')
                continue
            if upstream_overhang in existing_overhangs:
                upstream_overhang_obj = existing_overhangs[upstream_overhang]
            else:
                upstream_overhang_obj = Overhang.objects.create(
                    name = upstream_overhang,
                    sequence = upstream_overhang,
                    color = upstream_overhang_color
                    )
            if downstream_overhang in existing_overhangs:
                downstream_overhang_obj = existing_overhangs[downstream_overhang]
            else:
                downstream_overhang_obj = Overhang.objects.create(
                    name = downstream_overhang,
                    sequence = downstream_overhang,
                    color = downstream_overhang_color
                    )
            part_obj = Magic_pool_part_type.objects.create(
                name = part_name,
                description = description,
                upstream_overhang = upstream_overhang_obj,
                downstream_overhang = downstream_overhang_obj,
                )
            for item in partinfo:
                info_obj = Magic_pool_part_type_info.objects.create(
                    magic_pool_part_type = part_obj,
                    name = item[0],
                    description = item[1]
                )
            created_count += 1

    sheet = wb_obj.get_sheet_by_name('overview')
    xlsx_header = []
    existing_vectors = {item.name:item for item in Vector_type.objects.all()}
    current_vector_name = None
    current_description = ''
    magic_pool_parts = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            continue
        else:
            cell_a = str(row[0])
            cell_b = str(row[1])
            cell_c = str(row[2])
            if cell_a != '' and cell_a != 'None' and cell_a is not None:
                if current_vector_name is not None and current_vector_name not in existing_vectors:
                        vector_obj = Vector_type.objects.create(
                            name = current_vector_name,
                            description = current_description
                        )
                        for item_index, item in enumerate(magic_pool_parts):
                            try:
                                magic_pool_part_obj = Magic_pool_part_type.objects.get(name=item)
                                vector_part_obj = Vector_type_part.objects.create(
                                    vector_type = vector_obj,
                                    part_type = magic_pool_part_obj,
                                    order = item_index
                                )
                            except Magic_pool_part_type.DoesNotExist:
                                print(item, ' NOT FOUND IN MAGIC POOL PARTS')
                magic_pool_parts = []
                current_vector_name = cell_a
                current_description = cell_b
                magic_pool_parts.append(cell_c)
            elif cell_c != '' and cell_c != 'None' and cell_c is not None:
                magic_pool_parts.append(cell_c)
    if magic_pool_parts and current_vector_name is not None:
        if current_vector_name not in existing_vectors:
            vector_obj = Vector_type.objects.create(
                name = current_vector_name,
                description = current_description
            )
            for item_index, item in enumerate(magic_pool_parts):
                if Magic_pool_part_type.objects.filter(name=item).exists():
                    magic_pool_part_obj = Magic_pool_part_type.objects.get(name=item)
                    vector_part_obj = Vector_type_part.objects.create(
                        vector_type = vector_obj,
                        part_type = magic_pool_part_obj,
                        order = item_index
                    )
                else:
                    print(item, ' NOT FOUND IN MAGIC POOL PARTS')


def create_magic_pool(magic_pool_rows, existing_magic_pools):
    name = ''
    description = ''
    antibiotic_resistance = ''
    vector_type_name = ''
    plasmids = {}
    for row in magic_pool_rows:
        if row[0] == 'Magic_Pool_Name':
            name = row[1]
        elif row[0] == 'Description':
            description = row[1]
        elif row[0] == 'Antibiotic resistance':
            antibiotic_resistance = row[1]
        elif row[0] == 'Magic Pool Vector Type':
            vector_type_name = row[1]
        elif row[0] == 'Part_Vector':
            continue
        else:
            plasmids[row[1]] = (row[0], row[2])
    if name == '':
        return 0
    if name in existing_magic_pools:
        return 0
    magic_pool = Magic_pool.objects.create(
        name = name,
        description = description,
        antibiotic_resistance = antibiotic_resistance
        )
    if Vector_type.objects.filter(name=vector_type_name).exists():
        magic_pool.vector_type = Vector_type.objects.get(name=vector_type_name)
        magic_pool.save()
    else:
        print(vector_type_name, 'not found in vector types')
    # Update plasmids
    for plasmid_name in plasmids.keys():
        if plasmid_name == '' or plasmid_name is None:
            continue
        if Plasmid.objects.filter(name=plasmid_name).exists():
            plasmid = Plasmid.objects.get(name=plasmid_name)
            plasmid_designation = plasmids[plasmid_name][1]
            if plasmid.magic_pool_designation != plasmid_designation:
                print('Magic pool designation is not consistent:', plasmid_designation, plasmid.magic_pool_designation)
                raise ValueError('Update list of plasmids first')
            magic_pool_part_type = plasmids[plasmid_name][0]
            if plasmid.magic_pool_part.name != magic_pool_part_type:
                print('Magic pool part type is not consistent:', magic_pool_part_type, plasmid.magic_pool_part.name)
                raise ValueError('Update list of plasmids first')
            plasmid.magic_pools.add(magic_pool)
            plasmid.save()
        else:
            print(plasmid_name, 'not found in plasmids')
    return 1
            

def import_magic_pools(xlsx_path):
    xlsx_path = Path(xlsx_path)
    wb_obj = openpyxl.load_workbook(xlsx_path)
    sheet = wb_obj.get_sheet_by_name('magic_pool_summary')
    magic_pool_rows = []
    created_count = 0
    existing_magic_pools = {item.name:item for item in Magic_pool.objects.all()}
    for i, row in enumerate(sheet.iter_rows()):
        key = row[0].value
        if key == 'Magic_Pool_Name' and magic_pool_rows:
            created_count += create_magic_pool(magic_pool_rows, existing_magic_pools)
            magic_pool_rows = []
        magic_pool_rows.append((row[0].value, row[1].value, row[2].value))
    if magic_pool_rows:
        created_count += create_magic_pool(magic_pool_rows, existing_magic_pools)

    sheet = wb_obj.get_sheet_by_name('magic_pool_strainID')
    existing_magic_pools = {item.name:item for item in Magic_pool.objects.all()}
    for i, row in enumerate(sheet.iter_rows()):
        strain_amd = row[0].value
        magic_pool_name = row[2].value
        if magic_pool_name == 'Magic Pool Plasmid ID':
            continue
        if magic_pool_name == '' or magic_pool_name is None:
            continue
        if magic_pool_name not in existing_magic_pools:
            raise ValueError(magic_pool_name, 'magic pool from magic_pool_summary not found')
        if Strain.objects.filter(amd_number=strain_amd).exists():
            existing_magic_pools[magic_pool_name].strain = Strain.objects.get(amd_number=strain_amd)
            existing_magic_pools[magic_pool_name].save()
        else:
            print(strain_amd, 'strain from not magic_pool_summary found')
    print(created_count, 'new magic pools created')


def create_strain(amd_number, strain_data):
    species = ''
    description = ''
    plasmid = ''
    name = ''
    if 'Species' in strain_data:
        species = strain_data['Species']
    if 'Description' in strain_data:
        description = strain_data['Description']
    if 'Plasmid' in strain_data:
        plasmid = strain_data['Plasmid']
    if 'Strain' in strain_data:
        name = strain_data['Strain']
    strain_obj = Strain.objects.create(
        name = name,
        amd_number = amd_number,
        description = description,
        species = species,
        plasmid = plasmid
    )
    for key, value in strain_data.items():
        if key in ('Strain', 'Description', 'Species', 'Plasmid'):
            continue
        if key == '' or key is None:
            continue
        if value == '' or value is None:
            continue
        strain_info_obj = Strain_info.objects.create(
            strain = strain_obj,
            param = key,
            value = value
        )
    return 1


def import_strains_table(xlsx_path, overwrite_existing=False):
    print(xlsx_path)
    ret = []
    xlsx_path = Path(xlsx_path)
    wb_obj = openpyxl.load_workbook(xlsx_path, data_only=True)
    created_count = 0
    for sheet in wb_obj.worksheets:
        print('Working on ' + sheet.title)
        xlsx_header = []
        data_imported = defaultdict(dict)
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                xlsx_header = row
            else:
                amd_number = str(row[0])
                if amd_number is None or not amd_number.startswith('AMD'):
                    continue
                ret.append(amd_number)
                for j, cell in enumerate(row):
                    if j == 0:
                        continue
                    if cell != '' and cell != 'None' and cell is not None:
                        data_imported[amd_number][xlsx_header[j]] = str(cell)
        if str(xlsx_header[0]) != 'Strain':
            print('Skipping ' + sheet.title)
            continue
        for amd_number, strain_data in data_imported.items():
            if Strain.objects.filter(amd_number=amd_number).exists():
                strain = Strain.objects.get(amd_number=amd_number)
                if not overwrite_existing:
                    continue
                for key,value in strain_data.items():
                    if key == 'Strain':
                        if strain.name != value:
                            strain.name = value
                            strain.save()
                    elif key == 'Plasmid':
                        if strain.plasmid != value:
                            strain.plasmid = value
                            strain.save()
                    elif key == 'Description':
                        if strain.description != value:
                            strain.description = value
                            strain.save()
                    elif key == 'Species':
                        if strain.species != value:
                            strain.species = value
                            strain.save()
                    elif key == '' or key is None:
                        continue
                    else:
                        if Strain_info.objects.filter(strain=strain.id,param=key).exists():
                            strain_info_obj = Strain_info.objects.get(strain=strain.id,param=key)
                            if strain_info_obj.value != value:
                                strain_info_obj.value = value
                                strain_info_obj.save()
                        elif value != '' and not value is None:
                            Strain_info.objects.create(
                                strain = strain,
                                param = key,
                                value = value
                                )
            else:
                created_count += create_strain(amd_number, strain_data)
    print(created_count, 'new strains created')
    return ret


def create_oligo(name, oligo_data):
    sequence = ''
    description = ''
    if 'Sequence' in oligo_data:
        sequence = oligo_data['Sequence']
    if 'Purpose' in oligo_data:
        description = oligo_data['Purpose']
    oligo_obj = Oligo.objects.create(
        name = name,
        sequence = sequence,
        description = description
    )
    for key, value in oligo_data.items():
        if key in ('Sequence', 'Purpose'):
            continue
        if key == '' or key is None:
            continue
        if value == '' or value is None:
            continue
        oligo_info_obj = Oligo_info.objects.create(
            oligo = oligo_obj,
            param = key,
            value = value
        )
    return 1


def import_oligos_table(xlsx_path, overwrite_existing=False):
    print(xlsx_path)
    xlsx_path = Path(xlsx_path)
    wb_obj = openpyxl.load_workbook(xlsx_path, data_only=True)
    created_count = 0
    ret = []
    for sheet in wb_obj.worksheets:
        print('Working on ' + sheet.title)
        xlsx_header = []
        data_imported = defaultdict(dict)
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                xlsx_header = row
            else:
                name = str(row[0])
                if name is None or name == '':
                    continue
                ret.append(name)
                for j, cell in enumerate(row):
                    if j == 0:
                        continue
                    if cell != '' and cell != 'None' and cell is not None:
                        data_imported[name][xlsx_header[j]] = str(cell)
        for name, oligo_data in data_imported.items():
            if Oligo.objects.filter(name=name).exists():
                oligo = Oligo.objects.get(name=name)
                if not overwrite_existing:
                    continue
                for key,value in oligo_data.items():
                    if key == 'Sequence':
                        if oligo.sequence != value:
                            oligo.sequence = value
                            oligo.save()
                    elif key == 'Purpose':
                        if oligo.description != value:
                            oligo.description = value
                            oligo.save()
                    elif key == '' or key is None:
                        continue
                    else:
                        if Oligo_info.objects.filter(oligo=oligo.id,param=key).exists():
                            oligo_info_obj = Oligo_info.objects.get(oligo=oligo.id,param=key)
                            if oligo_info_obj.value != value:
                                oligo_info_obj.value = value
                                oligo_info_obj.save()
                        elif value != '' and not value is None:
                            Oligo_info.objects.create(
                                oligo = oligo,
                                param = key,
                                value = value
                                )
            else:
                print(oligo_data)
                created_count += create_oligo(name, oligo_data)
    print(created_count, 'new oligos created')
    return(ret)
    

def make_blast_databases():
    nucl_db_file = export_contigs()
    cmd = ['makeblastdb', '-dbtype', 'nucl', '-in',
           nucl_db_file, '-out',
           '/mnt/data/work/Plasmids/plasmidoro/data/blast_nucl'
           ]
    print(' '.join(cmd))
    with Popen(cmd, stdout=PIPE, bufsize=1, universal_newlines=True) as proc:
        for line in proc.stdout:
            print(line.rstrip('\n\r'))
    if proc.returncode != 0:
        # Suppress false positive no-member error
        # (see https://github.com/PyCQA/pylint/issues/1860)
        # pylint: disable=no-member
        raise CalledProcessError(proc.returncode, proc.args)
    
    prot_db_file = export_proteins()
    cmd = ['makeblastdb', '-dbtype', 'prot', '-in',
           prot_db_file, '-out',
           '/mnt/data/work/Plasmids/plasmidoro/data/blast_prot'
           ]
    print(' '.join(cmd))
    with Popen(cmd, stdout=PIPE, bufsize=1, universal_newlines=True) as proc:
        for line in proc.stdout:
            print(line.rstrip('\n\r'))
    if proc.returncode != 0:
        # Suppress false positive no-member error
        # (see https://github.com/PyCQA/pylint/issues/1860)
        # pylint: disable=no-member
        raise CalledProcessError(proc.returncode, proc.args)
    
    
def export_contigs():
    """
        Writes all nucleotide sequences into FASTA file
        for blast database generation
    """
    nucl_db_file = '/mnt/data/work/Plasmids/plasmidoro/data/nucl.fna'
    with open(nucl_db_file, 'w') as outfile:
        for item in Plasmid.objects.values('id', 'name', 'amd_number', 'sequence'):
            if item['sequence'] != '':
                outfile.write('>' + str(item['id']) + '|' + ''.join([i if ord(i) < 128 else '.' for i in item['name'].replace(' ', '_')]) + '|plasmid|' + item['amd_number'] +
                          '\n' + item['sequence'] +
                          '\n')
        for item in Oligo.objects.values('id', 'name', 'sequence'):
            if item['sequence'] != '':
                outfile.write('>' + str(item['id']) + '|' + ''.join([i if ord(i) < 128 else '.' for i in item['name'].replace(' ', '_')]) + '|oligo|' +
                          '\n' + item['sequence'] +
                          '\n')
    return nucl_db_file
    
    
def export_proteins():
    """
        Writes all protein sequences into FASTA file
        for blast database generation
    """
    prot_db_file = '/mnt/data/work/Plasmids/plasmidoro/data/prot.fna'
    with open(prot_db_file, 'w') as outfile:
        for protein in Protein.objects.all():
            if protein.sequence != '':
                protein_name = protein.name.replace('&nbsp;', '_')
                outfile.write('>' + str(protein.id) + '|' + ''.join([i if ord(i) < 128 else '.' for i in protein_name.replace(' ', '_')]) + '|' +
                    ''.join([i if ord(i) < 128 else '.' for i in protein.feature.plasmid.name.replace(' ', '_')]) + '|' + protein.feature.location_str +
                    '\n' + protein.sequence +
                    '\n')
    return prot_db_file
    
    

